from flask import Flask, render_template, jsonify
import requests
from datetime import datetime, timedelta
import os

app = Flask(__name__)

TRELLO_KEY   = os.environ.get('TRELLO_KEY',   '')
TRELLO_TOKEN = os.environ.get('TRELLO_TOKEN', '')
BOARD_ID     = 'BSMbxCEC'

AMO_TOKEN    = os.environ.get('AMO_TOKEN', '')
AMO_DOMAIN   = 'infogrowbccom.amocrm.ru'
AMO_PIPELINE = 7514314  # Продажи

META_TOKEN   = os.environ.get('META_TOKEN', '')
META_ACCOUNT = 'act_840654057640596'

RESIDENTS_SHEET_ID = '1FCRtmU9D9YeT9xHRAwEqiW5jgA-EcjyE7YBW3-E71qs'
PAYMENTS_SHEET_ID  = '1h4zh7mFTKyfGUWOsyk1us9EfP379qd2P'
PAYMENTS_GID       = '1169641494'

MEMBERSHIP_WEIGHTS = {
    'Резидент':          1.0,
    '1/2 Резидент':      0.5,
    '1/2 Резидент Women': 0.5,
}

# Квартальные цели по платящим резидентам
QUARTERLY_GOALS = {
    1: 45,   # Q1: январь–март
    2: 55,   # Q2: апрель–июнь
    3: 62,   # Q3: июль–сентябрь
    4: 72,   # Q4: октябрь–декабрь
}

# Посещаемость — ежемесячные файлы (сводная вкладка gid)
ATTENDANCE_SHEETS = {
    'Январь':  {'id': '1wWrJNqfCvmWecNHFOHQ8Gksft5SCQsogdDMy8d3VwIM', 'gid': '411480352'},
    'Февраль': {'id': '1XKCtDfw2obueBK3HJ9UXlc-5siKE4mphfiXYyixWHo8', 'gid': '411480352'},
    'Март':    {'id': '1R5q0WvnrsS6q8aucal3PGK8KQCNeVFHFQ9xENctgNPQ', 'gid': '411480352'},
    'Апрель':  {'id': '135D51Qubab4lGFy9Vp5a41-ziX_YlLC5nG-xE6_X2NA', 'gid': '411480352'},
    'Май':     {'id': '15__VbyyNTHXFWufA6pS_BOj6TtKSD2ERYnUUkm6xeck', 'gid': '411480352'},
    'Июнь':    {'id': '18TCohHjM2GPkn430uc4h0A0AGMRRYS8MUTT65JJiBDI', 'gid': '411480352'},
    'Июль':    {'id': '18uIkcV9pRs7-xvSFdMBc8KRQS5ymkmQ_PTnqFHehrnk', 'gid': '1364402557'},
}
ANNUAL_SHEET_ID = '1P-r6Q7uZ9aovbBFZqV17iHsRqTdT3m3BHxg6_BMY3Y8'
ANNUAL_GID      = '2092514904'
MONTH_NAMES_RU  = ['Январь','Февраль','Март','Апрель','Май','Июнь',
                   'Июль','Август','Сентябрь','Октябрь','Ноябрь','Декабрь']

BUDGET_SHEET_ID = '1JyvdruB8xc5YSVNQncFmT_aheaViO8d-roQ609SBL2s'
BUDGET_GID      = '1047075700'
MONTH_ROMAN     = {1:'I',2:'II',3:'III',4:'IV',5:'V',6:'VI',
                   7:'VII',8:'VIII',9:'IX',10:'X',11:'XI',12:'XII'}

TEAM = {
    'Даша':  {'work': 'Задачи Даша',    'backlog': 'Backlog_Даша'},
    'Алина': {'work': 'Алина_в работе', 'backlog': 'Backlog_Алина'},
    'Саша':  {'work': 'Саша_в работе',  'backlog': None},
    'Люба':  {'work': 'Задачи Люба',    'backlog': None},
}

# Системные статусы AmoCRM (закрытые сделки — есть в каждой воронке)
AMO_CLOSED_STATUSES = {142: 'Успешно реализовано', 143: 'Закрыто и не реализовано'}

def trello(path, **kw):
    r = requests.get(
        f'https://api.trello.com/1{path}',
        params={'key': TRELLO_KEY, 'token': TRELLO_TOKEN, **kw},
        timeout=10
    )
    r.raise_for_status()
    return r.json()

def amo(path, **params):
    r = requests.get(
        f'https://{AMO_DOMAIN}/api/v4{path}',
        headers={'Authorization': f'Bearer {AMO_TOKEN}'},
        params=params,
        timeout=10
    )
    r.raise_for_status()
    return r.json()

def get_week_range():
    today  = datetime.now()
    monday = today - timedelta(days=today.weekday())
    sunday = monday + timedelta(days=6)
    return (
        monday.replace(hour=0,  minute=0,  second=0,  microsecond=0),
        sunday.replace(hour=23, minute=59, second=59, microsecond=999999),
    )

# ── Trello ────────────────────────────────────────────────
def fetch_trello():
    lists = trello(
        f'/boards/{BOARD_ID}/lists',
        cards='open',
        card_fields='name,dueComplete,due'
    )
    lmap = {l['name']: l for l in lists}

    week_start, week_end = get_week_range()

    # Помощник: карточка попадает в эту неделю по дедлайну?
    def due_this_week(card):
        due = card.get('due')
        if not due:
            return False
        try:
            due_dt = datetime.fromisoformat(due.replace('Z', '+00:00')).replace(tzinfo=None)
            return week_start <= due_dt <= week_end
        except Exception:
            return False

    team = {}
    for name, cfg in TEAM.items():
        wl = lmap.get(cfg['work'], {})
        bl = lmap.get(cfg['backlog'] or '', {})
        wc = wl.get('cards', [])
        bc = bl.get('cards', [])

        # Статус по сотруднику — ВСЕ карточки в рабочей колонке
        done_cards   = [c for c in wc if c.get('dueComplete')]
        active_cards = [c for c in wc if not c.get('dueComplete')]

        team[name] = {
            'done':          [c['name'] for c in done_cards],
            'done_count':    len(done_cards),
            'work':          [c['name'] for c in active_cards],
            'work_count':    len(active_cards),
            'total_work':    len(wc),
            'backlog':       [c['name'] for c in bc],
            'backlog_count': len(bc),
        }

    # Прогресс недели — только карточки с дедлайном НА ЭТУ НЕДЕЛЮ
    week_work_cards = []
    for name, cfg in TEAM.items():
        wl = lmap.get(cfg['work'], {})
        for c in wl.get('cards', []):
            if due_this_week(c):
                week_work_cards.append(c)

    # Колонка "Неделя" — перенесённые завершённые
    week_done_cards = []
    week_list_name  = None
    for lname, ldata in lmap.items():
        if 'неделя' in lname.lower():
            week_done_cards += ldata.get('cards', [])
            week_list_name = lname

    # Подсчёт прогресса
    done_in_work    = [c for c in week_work_cards if c.get('dueComplete')]
    week_done_count = len(done_in_work) + len(week_done_cards)
    week_total      = len(week_work_cards) + len(week_done_cards)
    pct = int(week_done_count / week_total * 100) if week_total else 0

    week = {
        'name':      f'{week_start.strftime("%d.%m")} — {week_end.strftime("%d.%m")}',
        'done':      week_done_count,
        'planned':   week_total,
        'remain':    week_total - week_done_count,
        'pct':       pct,
        'cards':     [c['name'] for c in week_done_cards],
        'list_name': week_list_name,
    }

    return team, week

# ── AmoCRM ────────────────────────────────────────────────
def amo_get_all(path, **params):
    """Загружает все страницы из AmoCRM API"""
    all_items = []
    page = 1
    while True:
        try:
            data  = amo(path, page=page, limit=250, **params)
            items = data.get('_embedded', {}).get(path.strip('/').split('/')[-1], [])
            if not items:
                break
            all_items.extend(items)
            if len(items) < 250:
                break
            page += 1
        except Exception:
            break
    return all_items

def fetch_pipelines():
    """Загружает все воронки и их этапы, возвращает карту status_id -> info"""
    status_map = {}      # status_id -> {'name', 'pipeline_id', 'pipeline_name', 'sort'}
    pipeline_list = []   # [(pipeline_id, pipeline_name, sort)]

    page = 1
    while True:
        try:
            data      = amo('/leads/pipelines', page=page, limit=250)
            pipelines = data.get('_embedded', {}).get('pipelines', [])
            if not pipelines:
                break
            for p in pipelines:
                pid   = p['id']
                pname = p['name']
                psort = p.get('sort', 999)
                pipeline_list.append((pid, pname, psort))
                for s in p.get('_embedded', {}).get('statuses', []):
                    status_map[s['id']] = {
                        'name':          s['name'],
                        'pipeline_id':   pid,
                        'pipeline_name': pname,
                        'sort':          s.get('sort', 999),
                    }
            if len(pipelines) < 250:
                break
            page += 1
        except Exception:
            break

    # Добавляем системные статусы (закрытые), они не всегда приходят в списке
    for sid, sname in AMO_CLOSED_STATUSES.items():
        if sid not in status_map:
            status_map[sid] = {
                'name': sname, 'pipeline_id': None,
                'pipeline_name': 'Системные', 'sort': 9999,
            }

    pipeline_list.sort(key=lambda x: x[2])
    return status_map, pipeline_list


def fetch_crm():
    if not AMO_TOKEN:
        return None

    try:
        from collections import Counter, defaultdict

        now_ts      = int(datetime.now().timestamp())
        month_start = int(datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0).timestamp())

        # 1. Загружаем карту этапов
        status_map, _ = fetch_pipelines()

        CLOSED = {142, 143}

        # Имена этапов которые исключаем полностью
        EXCLUDE_STAGE_NAMES = {'архив'}

        # 2. Только лиды из воронки Продажи
        leads = amo_get_all('/leads', **{'filter[pipeline_id]': AMO_PIPELINE})

        # Активные: без системных закрытых и без этапов-исключений (Архив)
        def stage_name_of(lead):
            info = status_map.get(lead['status_id'])
            return (info['name'] if info else '').lower()

        active_leads = [
            l for l in leads
            if l['status_id'] not in CLOSED
            and stage_name_of(l) not in EXCLUDE_STAGE_NAMES
        ]
        active_total = len(active_leads)

        # 3. Новые за текущий месяц
        new_this_month = sum(
            1 for l in active_leads
            if l.get('created_at', 0) >= month_start
        )

        # 4. Воронка — только этапы Продажи
        counts_by_status = Counter(l['status_id'] for l in active_leads)

        prodazhi_stages = []
        for sid, info in sorted(
            ((k, v) for k, v in status_map.items() if v['pipeline_id'] == AMO_PIPELINE),
            key=lambda x: x[1]['sort']
        ):
            cnt = counts_by_status.get(sid, 0)
            if cnt > 0:
                prodazhi_stages.append({'name': info['name'], 'count': cnt, 'closed': False})

        if prodazhi_stages:
            max_cnt = max(s['count'] for s in prodazhi_stages)
            for s in prodazhi_stages:
                s['pct'] = int(s['count'] / max_cnt * 100)

        funnel_sections = [{
            'pipeline': 'Воронка продаж',
            'stages':   prodazhi_stages,
            'total':    active_total,
        }]

        # 5. Просроченные задачи — из активных лидов Продажи
        STUCK_EXCLUDE = {
            'резидент',
            'архив',
        }

        # Индекс активных лидов по ID
        active_lead_ids = {l['id'] for l in active_leads}

        # Загружаем просроченные незавершённые задачи
        overdue_tasks = []
        try:
            page = 1
            while True:
                data  = amo('/tasks', page=page, limit=250,
                            **{'filter[is_completed]': 0,
                               'filter[complete_till][to]': now_ts,
                               'filter[entity_type]': 'leads'})
                tasks = data.get('_embedded', {}).get('tasks', [])
                if not tasks:
                    break
                overdue_tasks.extend(tasks)
                if len(tasks) < 250:
                    break
                page += 1
        except Exception:
            pass

        stuck_by_stage = defaultdict(list)
        for task in overdue_tasks:
            lead_id = task.get('entity_id')
            if lead_id not in active_lead_ids:
                continue
            # Находим лид и его этап
            lead = next((l for l in active_leads if l['id'] == lead_id), None)
            if not lead:
                continue
            info       = status_map.get(lead['status_id'])
            stage_name = info['name'] if info else f'Этап {lead["status_id"]}'
            if stage_name.lower() in STUCK_EXCLUDE:
                continue
            days_overdue = (now_ts - task.get('complete_till', now_ts)) // 86400
            stuck_by_stage[stage_name].append(days_overdue)

        stuck_summary = []
        total_stuck   = 0
        for stage_name, days_list in sorted(stuck_by_stage.items(), key=lambda x: -len(x[1])):
            cnt      = len(days_list)
            total_stuck += cnt
            min_days, max_days = min(days_list), max(days_list)
            days_str = f'{min_days} дн.' if min_days == max_days else f'{min_days}–{max_days} дн.'
            stuck_summary.append({'stage': stage_name, 'count': cnt, 'days': days_str})

        return {
            'grand_total':     active_total,
            'active_total':    active_total,
            'new_month':       new_this_month,
            'stuck_count':     total_stuck,
            'stuck_summary':   stuck_summary,
            'funnel_sections': funnel_sections,
            'unknown_count':   0,
            'month_name':      ['январе','феврале','марте','апреле','мае','июне',
                                'июле','августе','сентябре','октябре','ноябре','декабре'][datetime.now().month-1],
        }
    except Exception as e:
        return {'error': str(e)}

# ── Residents ─────────────────────────────────────────────
def fetch_gsheet_csv(sheet_id, sheet_name=None, gid=None):
    import csv, io
    params = {'tqx': 'out:csv'}
    if sheet_name:
        params['sheet'] = sheet_name
    if gid:
        params['gid'] = gid
    r = requests.get(
        f'https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq',
        params=params,
        timeout=15
    )
    r.raise_for_status()
    reader = csv.DictReader(io.StringIO(r.text))
    return list(reader)

def fetch_residents():
    try:
        from collections import Counter

        # ── Резиденты ──
        rows = fetch_gsheet_csv(RESIDENTS_SHEET_ID, 'Резиденты')

        # Найдём колонку со статусом (содержит слово «статус»)
        status_col = next(
            (k for k in (rows[0].keys() if rows else []) if 'статус' in k.lower()),
            None
        )
        name_col = next(
            (k for k in (rows[0].keys() if rows else []) if 'имя' in k.lower() or 'фамили' in k.lower()),
            None
        )

        # Считаем резидентов, исключая «Вышел»
        active_rows = [r for r in rows if r.get(status_col, '').strip() != 'Вышел' and r.get(name_col, '').strip()]
        total = len(active_rows)

        # Разбивка по статусам
        status_counts = Counter(r.get(status_col, '').strip() for r in active_rows)

        # ── Дни рождения этой недели ──
        bday_rows = fetch_gsheet_csv(RESIDENTS_SHEET_ID, 'Д/Р')

        today      = datetime.now()
        week_start = today.replace(hour=0, minute=0, second=0, microsecond=0)
        week_end   = week_start + timedelta(days=7)

        # Находим колонку с именем гибко (может быть пробел в конце)
        bday_name_col = next(
            (k for k in (bday_rows[0].keys() if bday_rows else [])
             if 'фамили' in k.lower() or 'имя' in k.lower()),
            'Фамилия и Имя'
        )

        birthdays_week = []
        for b in bday_rows:
            name = b.get(bday_name_col, '').strip()
            if not name:
                continue

            # Пробуем «ДР в этом году» (DD.MM.YYYY), потом «День.Месяц» (DD.MM)
            dr_col = next((k for k in b.keys() if 'в этом году' in k.lower()), None)
            dm_col = next((k for k in b.keys() if 'день' in k.lower() and 'месяц' in k.lower()), None)

            raw = ''
            if dr_col:
                raw = b.get(dr_col, '').strip()
            if not raw and dm_col:
                raw = b.get(dm_col, '').strip()
            if not raw or '.' not in raw:
                continue

            try:
                parts = raw.replace('/', '.').split('.')
                day   = int(parts[0])
                month = int(parts[1])
                if day == 0 or month == 0:
                    continue
                bday = datetime(today.year, month, day)
                if week_start.date() <= bday.date() <= week_end.date():
                    months_ru = ['января','февраля','марта','апреля','мая','июня',
                                 'июля','августа','сентября','октября','ноября','декабря']
                    birthdays_week.append({
                        'name': name,
                        'day':  day,
                        'date': f'{day} {months_ru[month-1]}',
                    })
            except Exception:
                continue

        birthdays_week.sort(key=lambda x: x['day'])

        # ── Платящие резиденты (из таблицы оплат) ──
        pay_rows = fetch_gsheet_csv(PAYMENTS_SHEET_ID, gid=PAYMENTS_GID)

        total_paying  = 0.0
        total_paid_ok = 0.0

        # Находим колонки с именем/фамилией в таблице оплат
        pay_name_col    = None
        pay_surname_col = None
        pay_fio_col     = None
        if pay_rows:
            for k in pay_rows[0].keys():
                kl = k.lower()
                if 'фио' in kl or ('фамили' in kl and 'имя' in kl):
                    pay_fio_col = k
                elif 'фамили' in kl:
                    pay_surname_col = k
                elif 'имя' in kl and 'членство' not in kl:
                    pay_name_col = k

        today_dt = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        upcoming_payments = []

        for row in pay_rows:
            status     = row.get('Статус', '').strip()
            membership = row.get('Членство', '').strip()
            pay_status = row.get('Статус оплаты', '').strip().upper()

            if status != 'Active':
                continue

            weight = MEMBERSHIP_WEIGHTS.get(membership, 0)
            if weight == 0:
                continue

            total_paying += weight
            if pay_status == 'OK':
                total_paid_ok += weight

            # Проверяем скоро оплата / просрочка
            if pay_fio_col:
                name = row.get(pay_fio_col, '').strip()
            elif pay_surname_col or pay_name_col:
                surname = row.get(pay_surname_col, '').strip() if pay_surname_col else ''
                fname   = row.get(pay_name_col, '').strip() if pay_name_col else ''
                name    = f'{surname} {fname}'.strip()
            else:
                name = ''
            if not name:
                continue

            next_date_str = row.get('Следующая дата оплаты', '').strip()
            overdue_raw   = row.get('Дней просрочки', '').strip()

            overdue = 0
            try:
                overdue = int(float(overdue_raw)) if overdue_raw else 0
            except (ValueError, TypeError):
                overdue = 0

            days_until = None
            if next_date_str:
                for fmt_str in ('%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y'):
                    try:
                        nd = datetime.strptime(next_date_str, fmt_str)
                        days_until = (nd.replace(hour=0, minute=0, second=0, microsecond=0) - today_dt).days
                        break
                    except ValueError:
                        continue

            if overdue > 0 or (days_until is not None and days_until <= 30):
                upcoming_payments.append({
                    'name':       name,
                    'next_date':  next_date_str,
                    'overdue':    overdue,
                    'days_until': days_until,
                })

        upcoming_payments.sort(key=lambda x: (
            0 if x['overdue'] > 0 else 1,
            x['days_until'] if x['days_until'] is not None else 999
        ))

        def fmt(n):
            return int(n) if n == int(n) else n

        # ── KPI по кварталам ──
        current_q    = (datetime.now().month - 1) // 3 + 1
        current_goal = QUARTERLY_GOALS[current_q]
        gap          = current_goal - total_paid_ok
        pct_to_goal  = int(min(total_paid_ok / current_goal * 100, 100)) if current_goal else 0

        quarters = [
            {'q': q, 'name': name, 'goal': QUARTERLY_GOALS[q],
             'current': q == current_q}
            for q, name in [
                (1, 'Q1 · янв–мар'), (2, 'Q2 · апр–июн'),
                (3, 'Q3 · июл–сен'), (4, 'Q4 · окт–дек'),
            ]
        ]

        return {
            'total':              total,
            'status_counts':      dict(status_counts),
            'birthdays_week':     birthdays_week,
            'paying':             fmt(total_paying),
            'paid_ok':            fmt(total_paid_ok),
            'not_paid':           fmt(total_paying - total_paid_ok),
            'current_q':          current_q,
            'current_goal':       current_goal,
            'gap_to_goal':        fmt(gap) if gap > 0 else 0,
            'pct_to_goal':        pct_to_goal,
            'quarters':           quarters,
            'upcoming_payments':  upcoming_payments,
        }
    except Exception as e:
        return {'error': str(e)}

# ── Budget ────────────────────────────────────────────────
def parse_euro(s):
    """€2 495,00  →  2495.0"""
    s = str(s).replace('€','').replace('\xa0','').replace(' ','').replace(',','.').strip()
    try:
        return float(s)
    except ValueError:
        return 0.0

def fmt_euro(n):
    """2495.0  →  €2 495"""
    return '€' + f'{int(round(n)):,}'.replace(',', ' ')

def fetch_budget():
    try:
        import io as _io, openpyxl

        today        = datetime.now()
        curr_roman   = MONTH_ROMAN[today.month]
        target_month = today.month

        # Скачиваем XLSX — в нём и бюджет и календарь
        xlsx_r = requests.get(
            f'https://docs.google.com/spreadsheets/d/{BUDGET_SHEET_ID}/export?format=xlsx',
            timeout=30
        )
        xlsx_r.raise_for_status()
        wb = openpyxl.load_workbook(_io.BytesIO(xlsx_r.content), read_only=True, data_only=True)

        # ── Бюджет из «2026_Орг.расходы» ───────────────────────
        curr_plan = curr_fact = 0.0
        ws_bud = wb['2026_Орг.расходы']
        for row in ws_bud.iter_rows(values_only=True):
            if not row or len(row) < 5:
                continue
            col_a = str(row[0]).strip() if row[0] else ''
            col_c = str(row[2]).strip() if row[2] else ''
            if col_a == curr_roman and 'ИТОГО' in col_c:
                curr_plan = parse_euro(str(row[3]) if row[3] else '0')
                curr_fact = parse_euro(str(row[4]) if row[4] else '0')
                break

        curr_pct = int(curr_fact / curr_plan * 100) if curr_plan else 0

        # ── Кол-во опубликованных ивентов из «2026» ─────────────
        MONTH_MAP = {
            'ЯНВАРЬ':1,'ФЕВРАЛЬ':2,'МАРТ':3,'АПРЕЛЬ':4,'МАЙ':5,'ИЮНЬ':6,
            'ИЮЛЬ':7,'АВГУСТ':8,'СЕНТЯБРЬ':9,'ОКТЯБРЬ':10,'НОЯБРЬ':11,'ДЕКАБРЬ':12
        }
        READY = {'Готов к публикации', 'Готова к публикации'}
        published_count = 0
        curr_m = None
        ws_cal = wb['2026']

        # Месяц определяем ТОЛЬКО в строках-заголовках (col_a пустой или 'Месяц').
        # Это защита от ложных срабатываний: слово «МАЙ» входит в «МАСТЕРМАЙНД»,
        # что сбивало счётчик в середине июня.
        HEADER_COL_A = {'', 'none', 'месяц'}

        for row in ws_cal.iter_rows(values_only=True):
            col_a = str(row[0]).strip() if row and row[0] else ''

            # Ищем заголовок месяца только в «пустых» строках-разделителях
            if col_a.lower() in HEADER_COL_A:
                for cell in row:
                    if cell:
                        s = str(cell).upper().strip()
                        for mname, mnum in MONTH_MAP.items():
                            if s.startswith(mname):   # «ИЮНЬ» и «ИЮНЬ 2026» — ок; «МАСТЕРМАЙНД» — нет
                                curr_m = mnum

            if curr_m is not None and curr_m > target_month:
                break
            if curr_m != target_month:
                continue

            if col_a == 'Степень готовности':
                for val in row[1:]:
                    if val and str(val).strip() in READY:
                        published_count += 1

        return {
            'month':                 MONTH_NAMES_RU[today.month - 1],
            'curr_plan':             fmt_euro(curr_plan),
            'curr_fact':             fmt_euro(curr_fact),
            'curr_pct':              curr_pct,
            'published_event_count': published_count,
        }
    except Exception as e:
        return {'error': str(e)}

# ── Attendance ────────────────────────────────────────────
# Ключевые слова для исключения вкладок (падел, форум-группы)
ATT_SKIP_KEYWORDS = ['падел', 'форум-группа', ' фг', 'образец', 'сводная']

def fetch_attendance():
    try:
        import io as _io
        import csv as _csv
        import openpyxl
        from collections import defaultdict

        today           = datetime.now()
        curr_month_name = MONTH_NAMES_RU[today.month - 1]
        prev_month_name = MONTH_NAMES_RU[today.month - 2] if today.month > 1 else None

        sheet_info = ATTENDANCE_SHEETS.get(curr_month_name)
        if not sheet_info:
            return {'error': f'Нет данных для {curr_month_name}'}

        # ── 1. Скачиваем XLSX текущего месяца ──────────────────
        xlsx_r = requests.get(
            f'https://docs.google.com/spreadsheets/d/{sheet_info["id"]}/export?format=xlsx',
            timeout=30
        )
        xlsx_r.raise_for_status()
        wb = openpyxl.load_workbook(_io.BytesIO(xlsx_r.content), read_only=True, data_only=True)

        # ── 2. Парсим каждый ивент ─────────────────────────────
        # Основной список резидентов (для % от общего)
        res_rows   = fetch_gsheet_csv(RESIDENTS_SHEET_ID, 'Резиденты')
        status_col = next((k for k in (res_rows[0].keys() if res_rows else []) if 'статус' in k.lower()), None)
        name_col   = next((k for k in (res_rows[0].keys() if res_rows else []) if 'имя' in k.lower() or 'фамили' in k.lower()), None)

        active_res_names = {r.get(name_col,'').strip() for r in res_rows
                            if r.get(status_col,'').strip() != 'Вышел' and r.get(name_col,'').strip()}
        total_residents  = len(active_res_names)

        # Всего ивентов в месяце (без Сводной и Образца)
        SKIP_META = {'сводная', 'образец'}
        total_event_count = sum(
            1 for sh in wb.sheetnames
            if not any(kw in sh.lower() for kw in SKIP_META)
        )

        # Падел и баня — у них нет статуса в календаре, считаем отдельно
        PADEL_BANYA_KW = ['падел', 'баня']
        padel_banya_count = sum(
            1 for sh in wb.sheetnames
            if any(kw in sh.lower() for kw in PADEL_BANYA_KW)
        )

        events        = []
        no_show_count = defaultdict(int)   # имя → сколько раз зарегился и не пришёл

        for shname in wb.sheetnames:
            low = shname.lower()
            if any(kw in low for kw in ATT_SKIP_KEYWORDS):
                continue

            ws   = wb[shname]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue

            header = [str(c).strip().lower() if c else '' for c in rows[0]]
            try:
                reg_idx  = next(i for i,h in enumerate(header) if 'регистрация' in h)
                pres_idx = next(i for i,h in enumerate(header) if 'присутствие' in h)
            except StopIteration:
                continue

            registered = 0
            attended   = 0

            for row in rows[1:]:
                if not row or not row[0]:
                    continue
                name_val = str(row[0]).strip()
                reg_val  = str(row[reg_idx]).strip()  if row[reg_idx]  else ''
                pres_val = str(row[pres_idx]).strip() if row[pres_idx] else ''

                if reg_val in ('Да', 'Возможно'):
                    registered += 1
                    if pres_val != 'Да':
                        no_show_count[name_val] += 1
                if pres_val == 'Да':
                    attended += 1

            pct_of_total = round(attended / total_residents * 100) if total_residents else 0
            pct_of_reg   = round(attended / registered * 100) if registered else 0

            events.append({
                'name':          shname,
                'registered':    registered,
                'attended':      attended,
                'pct_total':     pct_of_total,
                'pct_reg':       pct_of_reg,
            })

        # Топ-5 «зарегистрировался и не пришёл»
        top_noshows = sorted(
            [{'name': n, 'count': c} for n, c in no_show_count.items() if c > 0],
            key=lambda x: -x['count']
        )[:5]

        # Средняя явка и лучший ивент (только у ивентов с посещаемостью)
        done_events = [e for e in events if e['attended'] > 0]
        avg_pct     = round(sum(e['pct_reg'] for e in done_events) / len(done_events)) if done_events else 0
        top_event   = max(done_events, key=lambda x: x['attended'])['name'].split(' ', 1)[1] if done_events else '—'

        # ── 3. Статус резидентов из Сводной ───────────────────
        summary_rows = fetch_gsheet_csv(sheet_info['id'], gid=sheet_info['gid'])
        monthly_data = {}
        for row in summary_rows:
            n = row.get('ФИО','').strip()
            if not n:
                continue
            try:
                p = int(row.get('Присутствие','0') or 0)
            except (ValueError, TypeError):
                p = 0
            monthly_data[n] = {
                'presence': p,
                'tariff':   row.get('Тариф','').strip(),
            }

        # Предыдущий месяц (только для определения «Новый»)
        prev_info = ATTENDANCE_SHEETS.get(prev_month_name) if prev_month_name else None
        prev_names = set()
        if prev_info:
            prev_rows = fetch_gsheet_csv(prev_info['id'], gid=prev_info['gid'])
            prev_names = {r.get('ФИО','').strip() for r in prev_rows if r.get('ФИО','').strip()}

        # Статус строится по основному списку резидентов (исключаем Вышедших)
        status_counts = defaultdict(int)
        residents_out = []

        for name in sorted(active_res_names):
            data   = monthly_data.get(name, {})
            tariff = data.get('tariff', '') or 'Резидент'
            if tariff == 'Deactive':
                continue
            p_curr = data.get('presence', 0)

            if tariff == 'Амбассадор':
                status = 'Амбассадор'
            elif name not in prev_names:
                status = 'Новый'
            elif p_curr >= 3:
                status = 'Активный'
            elif p_curr >= 1:
                status = 'Выпал'
            else:
                status = 'Под риском'

            status_counts[status] += 1
            residents_out.append({
                'name':   name,
                'tariff': tariff,
                'p_curr': p_curr,
                'status': status,
            })

        STATUS_ORDER = {'Активный':0,'Выпал':1,'Под риском':2,'Новый':3,'Амбассадор':4}
        residents_out.sort(key=lambda x: (STATUS_ORDER.get(x['status'],9), x['name']))

        # ── 4. Годовая статистика ──────────────────────────────
        def raw_csv(sheet_id, gid):
            r = requests.get(
                f'https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq',
                params={'tqx': 'out:csv', 'gid': gid}, timeout=15
            )
            r.raise_for_status()
            return list(_csv.reader(_io.StringIO(r.text)))

        annual_rows = raw_csv(ANNUAL_SHEET_ID, ANNUAL_GID)
        ann_reg = ann_conf = ann_pres = 0
        annual_top = []

        for row in annual_rows[1:]:
            if len(row) < 5:
                continue
            n = row[1].strip()
            if not n:
                continue
            try:
                reg  = int(row[2]) if row[2].strip() else 0
                conf = int(row[3]) if row[3].strip() else 0
                pres = int(row[4]) if row[4].strip() else 0
            except (ValueError, IndexError):
                continue
            ann_reg  += reg
            ann_conf += conf
            ann_pres += pres
            if pres > 0:
                annual_top.append({'name': n, 'reg': reg, 'conf': conf, 'pres': pres})

        annual_top.sort(key=lambda x: -x['pres'])

        return {
            'curr_month':        curr_month_name,
            'prev_month':        prev_month_name or '—',
            'total_residents':   total_residents,
            'events':            events,
            'event_count':       len(done_events),
            'total_event_count':   total_event_count,
            'padel_banya_count':   padel_banya_count,
            'avg_pct':           avg_pct,
            'top_event':         top_event,
            'top_noshows':     top_noshows,
            'status_counts':   dict(status_counts),
            'residents':       residents_out,
            'annual_reg':      ann_reg,
            'annual_conf':     ann_conf,
            'annual_pres':     ann_pres,
            'annual_top':      annual_top[:10],
        }
    except Exception as e:
        import traceback
        return {'error': str(e), 'trace': traceback.format_exc()}

# ── Meta Ads ──────────────────────────────────────────────
def fetch_meta():
    try:
        if not META_TOKEN:
            return {'error': 'META_TOKEN не задан'}

        base = 'https://graph.facebook.com/v19.0'
        fields = 'impressions,clicks,spend,cpm,cpc,actions,cost_per_action_type'

        # Статистика за текущий месяц
        r = requests.get(
            f'{base}/{META_ACCOUNT}/insights',
            params={
                'fields':      fields,
                'date_preset': 'this_month',
                'access_token': META_TOKEN,
            },
            timeout=15
        )
        r.raise_for_status()
        data = r.json().get('data', [])
        if not data:
            return {'error': 'Нет данных от Meta'}

        d = data[0]
        spend      = float(d.get('spend', 0))
        impressions = int(d.get('impressions', 0))
        clicks     = int(d.get('clicks', 0))
        cpm        = float(d.get('cpm', 0))
        cpc        = float(d.get('cpc', 0))

        # Лиды из actions
        actions = d.get('actions', [])
        leads = next((int(a['value']) for a in actions if a['action_type'] == 'lead'), 0)

        cpl = round(spend / leads, 2) if leads else 0

        # Статистика по кампаниям
        rc = requests.get(
            f'{base}/{META_ACCOUNT}/insights',
            params={
                'fields':      'campaign_name,impressions,clicks,spend,actions',
                'date_preset': 'this_month',
                'level':       'campaign',
                'access_token': META_TOKEN,
            },
            timeout=15
        )
        rc.raise_for_status()
        campaigns_raw = rc.json().get('data', [])

        campaigns = []
        for c in campaigns_raw:
            c_spend  = float(c.get('spend', 0))
            c_actions = c.get('actions', [])
            c_leads  = next((int(a['value']) for a in c_actions if a['action_type'] == 'lead'), 0)
            c_cpl    = round(c_spend / c_leads, 2) if c_leads else None
            campaigns.append({
                'name':    c.get('campaign_name', ''),
                'spend':   round(c_spend, 2),
                'leads':   c_leads,
                'clicks':  int(c.get('clicks', 0)),
                'cpl':     c_cpl,
            })

        # Лучший CPL среди кампаний с лидами
        with_leads = [c for c in campaigns if c['leads'] > 0]
        best_cpl = min(with_leads, key=lambda x: x['cpl'])['cpl'] if with_leads else None

        # Лиды из AmoCRM за текущий месяц
        amo_leads = None
        amo_cpl   = None
        if AMO_TOKEN:
            try:
                month_start = int(datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0).timestamp())
                all_leads   = amo_get_all('/leads', **{'filter[pipeline_id]': AMO_PIPELINE})
                # Считаем ВСЕ лиды за месяц включая нерелевантные — для сравнения с Meta
                amo_leads   = sum(
                    1 for l in all_leads
                    if l.get('created_at', 0) >= month_start
                )
                amo_cpl = round(spend / amo_leads, 2) if amo_leads else None
            except Exception:
                pass

        today = datetime.now()
        return {
            'month':       MONTH_NAMES_RU[today.month - 1],
            'spend':       round(spend, 2),
            'impressions': impressions,
            'clicks':      clicks,
            'leads':       leads,
            'cpl':         cpl,
            'best_cpl':    best_cpl,
            'cpm':         round(cpm, 2),
            'cpc':         round(cpc, 2),
            'campaigns':   sorted(campaigns, key=lambda x: -x['leads']),
            'amo_leads':   amo_leads,
            'amo_cpl':     amo_cpl,
        }
    except Exception as e:
        return {'error': str(e)}

# ── Routes ────────────────────────────────────────────────
@app.route('/')
def index():
    try:
        team, week = fetch_trello()
        crm        = fetch_crm()
        residents  = fetch_residents()
        attendance = fetch_attendance()
        budget     = fetch_budget()
        meta       = fetch_meta()
        error      = None
    except Exception as e:
        team, week, crm, residents, attendance, budget, meta = {}, {}, None, None, None, None, None
        error = str(e)
    return render_template('index.html',
        team=team, week=week, crm=crm, residents=residents,
        attendance=attendance, budget=budget, meta=meta,
        error=error,
        updated=datetime.now().strftime('%d.%m.%Y в %H:%M')
    )

@app.route('/debug-bdays')
def debug_bdays():
    try:
        from datetime import datetime, timedelta
        rows  = fetch_gsheet_csv(RESIDENTS_SHEET_ID, 'Д/Р')
        if not rows:
            return 'Лист Д/Р пустой или не найден'

        today      = datetime.now()
        week_start = today.replace(hour=0, minute=0, second=0, microsecond=0)
        week_end   = week_start + timedelta(days=7)

        headers = list(rows[0].keys())
        dr_col  = next((k for k in headers if 'в этом году' in k.lower()), None)
        dm_col  = next((k for k in headers if 'день' in k.lower() and 'месяц' in k.lower()), None)

        lines = [
            f'Сегодня: {today.date()}',
            f'Период: {week_start.date()} — {week_end.date()}',
            f'Колонки: {headers}',
            f'dr_col (ДР в этом году): {dr_col}',
            f'dm_col (День.Месяц): {dm_col}',
            '',
            '=== Июльские строки ===',
        ]

        for b in rows:
            name = b.get('Фамилия и Имя', '').strip()
            raw  = b.get(dr_col, '').strip() if dr_col else ''
            if not raw:
                raw = b.get(dm_col, '').strip() if dm_col else ''
            month_txt = b.get('Месяц (текст)', '').strip().lower()
            if 'июл' not in month_txt and '07' not in raw:
                continue
            lines.append(f'{name} | raw={repr(raw)} | месяц={month_txt}')
            try:
                parts = raw.replace('/', '.').split('.')
                day   = int(parts[0])
                month = int(parts[1])
                bday  = datetime(today.year, month, day)
                inrange = week_start.date() <= bday.date() <= week_end.date()
                lines.append(f'  → day={day} month={month} bday={bday.date()} in_range={inrange}')
            except Exception as ex:
                lines.append(f'  → ОШИБКА ПАРСИНГА: {ex}')

        return '<pre>' + '\n'.join(lines) + '</pre>'
    except Exception as e:
        import traceback
        return f'<pre>Ошибка: {e}\n{traceback.format_exc()}</pre>'

@app.route('/debug-payments')
def debug_payments():
    try:
        rows = fetch_gsheet_csv(PAYMENTS_SHEET_ID, gid=PAYMENTS_GID)
        if not rows:
            return '<pre>Таблица пустая</pre>'

        headers = list(rows[0].keys())
        today_dt = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

        lines = [
            f'Колонки: {headers}',
            f'Сегодня: {today_dt.date()}',
            f'Всего строк: {len(rows)}',
            '',
            '=== Active резиденты (первые 10) ===',
        ]

        active = [r for r in rows if r.get('Статус', '').strip() == 'Active']
        for r in active[:10]:
            nd  = r.get('Следующая дата оплаты', 'НЕТ КОЛОНКИ')
            od  = r.get('Дней просрочки', 'НЕТ КОЛОНКИ')
            mem = r.get('Членство', '')
            lines.append(f'  Членство={mem!r} | Следующая дата={nd!r} | Дней просрочки={od!r}')

        lines += ['', f'Active строк: {len(active)}']
        return '<pre>' + '\n'.join(lines) + '</pre>'
    except Exception as e:
        import traceback
        return f'<pre>Ошибка: {e}\n{traceback.format_exc()}</pre>'

@app.route('/api')
def api():
    try:
        team, week = fetch_trello()
        crm        = fetch_crm()
        return jsonify({'team': team, 'week': week, 'crm': crm})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)), debug=False)
